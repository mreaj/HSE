"""Document ingestion: parse -> chunk -> embed -> upsert, with an on-disk registry.

Incremental: files whose SharePoint eTag hasn't changed are skipped on re-sync.
Deleting a document also removes its vectors.
"""
import os
import re
import io
import json

from pypdf import PdfReader
import docx as _docx
from openpyxl import load_workbook

from config import load_config
import graph_client as gc
import vectorstore as vs

REGISTRY_PATH = os.path.expanduser(os.path.join("~", "hse_rag_registry.json"))

SKIP_EXT = (".mp4", ".mov", ".avi", ".mkv", ".wmv", ".mp3", ".wav",
            ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".svg", ".zip", ".7z", ".exe")


def load_registry():
    if os.path.exists(REGISTRY_PATH):
        with open(REGISTRY_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_registry(reg):
    with open(REGISTRY_PATH, "w", encoding="utf-8") as f:
        json.dump(reg, f, indent=2)


REGISTRY = load_registry()


def extract_text(name, data):
    ext = name.lower().rsplit(".", 1)[-1] if "." in name else ""
    text = ""
    try:
        if ext == "pdf":
            reader = PdfReader(io.BytesIO(data))
            text = "\n".join((p.extract_text() or "") for p in reader.pages)
        elif ext == "docx":
            d = _docx.Document(io.BytesIO(data))
            text = "\n".join(p.text for p in d.paragraphs)
        elif ext in ("xlsx", "xlsm"):
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            rows = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    rows.append(" ".join("" if c is None else str(c) for c in row))
            text = "\n".join(rows)
        elif ext in ("txt", "md", "csv", "log"):
            text = data.decode("utf-8", errors="replace")
        elif ext in ("html", "htm"):
            from bs4 import BeautifulSoup
            soup = BeautifulSoup(data, "html.parser")
            for s in soup(["script", "style", "nav", "footer"]):
                s.extract()
            text = soup.get_text(" ")
    except Exception as e:
        print("  parse error:", name, e)
    return text or ""


def chunk_text(text):
    cfg = load_config()
    text = re.sub(r"\s+\n", "\n", text).strip()
    size, overlap = cfg["CHUNK_SIZE"], cfg["CHUNK_OVERLAP"]
    chunks, i = [], 0
    while i < len(text):
        ch = text[i:i + size]
        if len(ch.strip()) > 50:
            chunks.append(ch)
        i += size - overlap
    return chunks


def _index_bytes(item_id, name, web_url, drive_id, access, data, source, etag=None, log=print):
    text = extract_text(name, data)
    chunks = chunk_text(text)
    if not chunks:
        log(f"  skip (no usable text): {name}")
        REGISTRY[item_id] = {"name": name, "web_url": web_url, "drive_id": drive_id,
                             "access": access, "n_chunks": 0, "source": source, "etag": etag}
        save_registry(REGISTRY)
        return 0
    vecs = []
    for i in range(0, len(chunks), 16):
        vecs.extend(vs.embed(chunks[i:i + 16]))
    points = [vs.make_point(v, {
                "item_id": item_id, "name": name, "web_url": web_url,
                "drive_id": drive_id, "access": access, "chunk": ch, "ord": idx})
              for idx, (ch, v) in enumerate(zip(chunks, vecs))]
    vs.delete_item(item_id)      # remove any previous version
    vs.upsert(points)
    REGISTRY[item_id] = {"name": name, "web_url": web_url, "drive_id": drive_id,
                         "access": access, "n_chunks": len(chunks), "source": source, "etag": etag}
    save_registry(REGISTRY)
    log(f"  indexed: {name} ({len(chunks)} chunks)")
    return len(chunks)


def index_document(site_id, drive_id, item, site_url, log=print):
    name = item["name"]
    if name.lower().endswith(SKIP_EXT):
        return 0
    cfg = load_config()
    if (item.get("size", 0) or 0) > cfg["MAX_FILE_MB"] * 1024 * 1024:
        log(f"  skip (>{cfg['MAX_FILE_MB']}MB): {name}")
        return 0
    etag = item.get("eTag")
    prev = REGISTRY.get(item["id"])
    if prev and prev.get("etag") == etag and prev.get("n_chunks", 0) > 0:
        return 0  # unchanged
    data = gc.download_drive_item(drive_id, item["id"], gc.get_app_token())
    return _index_bytes(item["id"], name, item.get("webUrl", ""), drive_id,
                        "graph", data, source=site_url, etag=etag, log=log)


def index_by_url(url, log=print):
    url = url.strip()
    tok = gc.get_app_token()
    try:
        item = gc.resolve_share(url, tok)
        drive_id = item["parentReference"]["driveId"]
        data = gc.download_drive_item(drive_id, item["id"], tok)
        log("  SharePoint/OneDrive document (access-checked)")
        return _index_bytes(item["id"], item["name"], item.get("webUrl", url),
                            drive_id, "graph", data, source=url,
                            etag=item.get("eTag"), log=log)
    except Exception as e:
        log(f"  not a SharePoint link ({str(e)[:60]}); fetching as public URL...")
    data, name = gc.http_fetch(url)
    return _index_bytes("url:" + url, name, url, None, "public", data, source=url, log=log)


def sync_site(site_url, libraries=None, log=print):
    """Crawl one site. libraries=None -> all libraries; else a list of library names."""
    vs.ensure_collection()
    cfg = load_config()
    tok = gc.get_app_token()
    site = gc.resolve_site(site_url, tok)
    all_drives = gc.list_drives(site["id"], tok)
    log("Libraries: " + ", ".join(d["name"] for d in all_drives))
    scanned = indexed = total_chunks = 0
    for drv in all_drives:
        if libraries and drv["name"] not in libraries:
            continue
        log(f"Library: {drv['name']}")
        for it in gc.iter_drive_files(site["id"], drv["id"], tok):
            scanned += 1
            try:
                c = index_document(site["id"], drv["id"], it, site_url, log)
                if c:
                    indexed += 1
                    total_chunks += c
                    if indexed % cfg["CHECKPOINT_EVERY"] == 0:
                        log(f"  checkpoint backup at {indexed} docs...")
                        vs.backup_to_sharepoint()
            except Exception as e:
                log(f"  FAILED {it['name']}: {e}")
            if scanned % 200 == 0:
                log(f"  ...{scanned} scanned, {indexed} indexed")
    vs.backup_to_sharepoint()
    log(f"Done. Scanned {scanned}, indexed/updated {indexed} docs, {total_chunks} chunks.")
    return total_chunks


def delete_document(item_id, log=print):
    vs.delete_item(item_id)
    REGISTRY.pop(item_id, None)
    save_registry(REGISTRY)
    vs.backup_to_sharepoint()
    log(f"Deleted document + vectors: {item_id}")
